from pandas import DataFrame, ExcelWriter
import openpyxl

class Pandawriter:
	"""
	Class writer takes care of writing the collected infromation from the scd file into the xls file
	"""
	scd = None
	list_of_IEDs = []
	datascd = None
	dataIED = DataFrame()
	EWriter = None

	@staticmethod
	def load(scd, list_of_IEDs):
		"""
		Static method load accepts the scd content object and list of processed IEDS
		:param scd: scd content processed object to be written
		:param list_of_IEDs: list of processed IEDs to be written
		"""
		Pandawriter.datascd = DataFrame(scd.values, index=[0])

		Pandawriter.list_of_IEDs = list_of_IEDs
		for IED in list_of_IEDs:
			df = DataFrame(IED.values, index=[0])
			Pandawriter.dataIED = Pandawriter.dataIED.append(df, ignore_index=True)

	@staticmethod
	def write(path_file):
		"""
		Method creates and ExcelWriter class and then writes the information loaded into the file at path

		:param path: file path
		"""

		# Pandawriter.datascd.to_excel(path_file, sheet_name='SCD META INFO')

		EWriter = ExcelWriter(path_file, engine='openpyxl')
		workbook = EWriter.book
		#worksheet = workbook.add.worksheet('FILE INFO')
		#EWriter.sheets['FILE INFO'] = worksheet
		Pandawriter.datascd.to_excel(EWriter, sheet_name="FILE INFO", startrow=0, startcol=0)

		Pandawriter.dataIED.to_excel(EWriter, sheet_name="FILE INFO", startrow=5, startcol=0)

		for IED in Pandawriter.list_of_IEDs:
			DataFrame().to_excel(EWriter, sheet_name=IED.values['name'])

		EWriter.save()


	def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
						   truncate_sheet=False,
						   **to_excel_kwargs):
		"""
		Append a DataFrame [df] to existing Excel file [filename]
		into [sheet_name] Sheet.
		If [filename] doesn't exist, then this function will create it.

		Parameters:
		  filename : File path or existing ExcelWriter
					 (Example: '/path/to/file.xlsx')
		  df : dataframe to save to workbook
		  sheet_name : Name of sheet which will contain DataFrame.
					   (default: 'Sheet1')
		  startrow : upper left cell row to dump data frame.
					 Per default (startrow=None) calculate the last row
					 in the existing DF and write to the next row...
		  truncate_sheet : truncate (remove and recreate) [sheet_name]
						   before writing DataFrame to Excel file
		  to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
							[can be dictionary]

		Returns: None
		"""
		from openpyxl import load_workbook

		import pandas as pd

		# ignore [engine] parameter if it was passed
		if 'engine' in to_excel_kwargs:
			to_excel_kwargs.pop('engine')

		writer = pd.ExcelWriter(filename, engine='openpyxl')

		# Python 2.x: define [FileNotFoundError] exception if it doesn't exist
		try:
			FileNotFoundError
		except NameError:
			FileNotFoundError = IOError

		try:
			# try to open an existing workbook
			writer.book = load_workbook(filename)

			# get the last row in the existing Excel sheet
			# if it was not specified explicitly
			if startrow is None and sheet_name in writer.book.sheetnames:
				startrow = writer.book[sheet_name].max_row

			# truncate sheet
			if truncate_sheet and sheet_name in writer.book.sheetnames:
				# index of [sheet_name] sheet
				idx = writer.book.sheetnames.index(sheet_name)
				# remove [sheet_name]
				writer.book.remove(writer.book.worksheets[idx])
				# create an empty sheet [sheet_name] using old index
				writer.book.create_sheet(sheet_name, idx)

			# copy existing sheets
			writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
		except FileNotFoundError:
			# file does not exist yet, we will create it
			pass

		if startrow is None:
			startrow = 0

		# write out the new sheet
		df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

		# save the workbook
		writer.save()